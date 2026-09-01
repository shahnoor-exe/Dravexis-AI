"""
artifact_compiler/xlsx_compiler.py — Openpyxl spreadsheet generator.

Generates an .xlsx with:
  - Visible Excel formulas (not hardcoded values)
  - Source/evidence metadata sheet
  - Synthetic calculation section with corrosion remaining-life formula
  - Provenance sheet
  - Disclaimer in every sheet

All outputs go to data/artifacts/ only.
"""
from __future__ import annotations

import logging
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..config import BASE_DIR

logger = logging.getLogger(__name__)

DISCLAIMER = (
    "Prototype output — not an engineering approval or statutory determination."
)
ARTIFACT_DIR = BASE_DIR / "data" / "artifacts"
COMPILER_VERSION = f"xlsx_compiler-0.1.0 / openpyxl-{openpyxl.__version__}"

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
_WARN_FONT = Font(color="C00000", bold=True, size=10)
_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    cell.border = _BORDER


def _cell(ws, row: int, col: int, value, bold=False, warn=False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _BORDER
    if bold:
        cell.font = Font(bold=True)
    if warn:
        cell.fill = _WARN_FILL
        cell.font = _WARN_FONT


def generate(
    *,
    query: str,
    evidence: list[dict],
    session_id: str,
    model_role: str | None,
    label: str = "SYNTHETIC",
    request_id: str | None = None,
) -> dict[str, Any]:
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    request_id = request_id or str(uuid.uuid4())
    ts = datetime.now(timezone.utc)
    safe_name = f"analysis_{request_id[:8]}_{ts.strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = ARTIFACT_DIR / safe_name

    try:
        out_path.resolve().relative_to(ARTIFACT_DIR.resolve())
    except ValueError:
        return {"status": "error", "error": "path traversal rejected"}

    try:
        wb = openpyxl.Workbook()

        # ------------------------------------------------------------------ #
        # Sheet 1: Summary                                                     #
        # ------------------------------------------------------------------ #
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.column_dimensions["A"].width = 28
        ws_sum.column_dimensions["B"].width = 60

        ws_sum["A1"] = "MRPL Sovereign AI Workbench — Demo Analysis"
        ws_sum["A1"].font = Font(bold=True, size=14, color="1F3864")
        ws_sum.merge_cells("A1:B1")

        ws_sum["A2"] = DISCLAIMER
        ws_sum["A2"].font = _WARN_FONT
        ws_sum["A2"].fill = _WARN_FILL
        ws_sum.merge_cells("A2:B2")

        meta = [
            ("Document ID", request_id),
            ("Generated (UTC)", ts.strftime("%Y-%m-%d %H:%M:%S")),
            ("Session ID", session_id),
            ("Query", query[:200]),
            ("Data Label", label),
            ("Model Role", model_role or "N/A"),
            ("Compiler", COMPILER_VERSION),
            ("Vision Status", "VISION_UNAVAILABLE"),
            ("Sandbox Mode", "DEGRADED_SANDBOX"),
        ]
        for i, (k, v) in enumerate(meta, start=4):
            _cell(ws_sum, i, 1, k, bold=True)
            _cell(ws_sum, i, 2, v, warn=(v in ("VISION_UNAVAILABLE", "DEGRADED_SANDBOX")))

        # ------------------------------------------------------------------ #
        # Sheet 2: Calculation (visible formulas)                              #
        # ------------------------------------------------------------------ #
        ws_calc = wb.create_sheet("Calculation")
        ws_calc.column_dimensions["A"].width = 32
        ws_calc.column_dimensions["B"].width = 20
        ws_calc.column_dimensions["C"].width = 20
        ws_calc.column_dimensions["D"].width = 40

        ws_calc["A1"] = "Synthetic Corrosion Remaining-Life Calculation"
        ws_calc["A1"].font = Font(bold=True, size=13, color="1F3864")
        ws_calc.merge_cells("A1:D1")

        ws_calc["A2"] = DISCLAIMER
        ws_calc["A2"].font = _WARN_FONT
        ws_calc.merge_cells("A2:D2")

        _header(ws_calc, 4, 1, "Parameter")
        _header(ws_calc, 4, 2, "Value")
        _header(ws_calc, 4, 3, "Unit")
        _header(ws_calc, 4, 4, "Note")

        inputs = [
            ("Measured Thickness", 8.5, "mm", "Synthetic demo value"),
            ("Min Required Thickness", 6.0, "mm", "ASME B31.3 design minimum"),
            ("Corrosion Rate", 0.30, "mm/year", "Synthetic demo value"),
        ]
        for row_i, (param, val, unit, note) in enumerate(inputs, start=5):
            _cell(ws_calc, row_i, 1, param, bold=True)
            _cell(ws_calc, row_i, 2, val)
            _cell(ws_calc, row_i, 3, unit)
            _cell(ws_calc, row_i, 4, note)

        # Remaining life = (B5 - B6) / B7 — visible formula
        _cell(ws_calc, 9, 1, "Remaining Life", bold=True)
        ws_calc["B9"] = "=(B5-B6)/B7"
        ws_calc["B9"].number_format = "0.00"
        _cell(ws_calc, 9, 3, "years")
        _cell(ws_calc, 9, 4, "Formula: (Measured - Min) / Corrosion Rate")

        ws_calc["A11"] = "⚠ Synthetic values only. Not derived from real inspection data."
        ws_calc["A11"].font = _WARN_FONT

        # ------------------------------------------------------------------ #
        # Sheet 3: Evidence                                                    #
        # ------------------------------------------------------------------ #
        ws_ev = wb.create_sheet("Evidence")
        ws_ev.column_dimensions["A"].width = 30
        ws_ev.column_dimensions["B"].width = 12
        ws_ev.column_dimensions["C"].width = 70

        _header(ws_ev, 1, 1, "Document ID")
        _header(ws_ev, 1, 2, "Score")
        _header(ws_ev, 1, 3, "Text Preview")

        if evidence:
            for row_i, ev in enumerate(evidence[:10], start=2):
                _cell(ws_ev, row_i, 1, ev.get("doc_id", ""))
                _cell(ws_ev, row_i, 2, round(ev.get("score", 0), 4))
                _cell(ws_ev, row_i, 3, ev.get("text_preview", "")[:300])
        else:
            ws_ev["A2"] = "No evidence retrieved"
            ws_ev["A2"].font = Font(italic=True)

        # ------------------------------------------------------------------ #
        # Sheet 4: Provenance                                                  #
        # ------------------------------------------------------------------ #
        ws_prov = wb.create_sheet("Provenance")
        ws_prov.column_dimensions["A"].width = 28
        ws_prov.column_dimensions["B"].width = 60
        prov_data = [
            ("request_id", request_id),
            ("session_id", session_id),
            ("timestamp_utc", ts.isoformat()),
            ("label", label),
            ("model_role", model_role or "N/A"),
            ("compiler", COMPILER_VERSION),
            ("evidence_count", str(len(evidence))),
            ("disclaimer", DISCLAIMER),
        ]
        for i, (k, v) in enumerate(prov_data, start=1):
            _cell(ws_prov, i, 1, k, bold=True)
            _cell(ws_prov, i, 2, v)

        wb.save(str(out_path))
        logger.info("XLSX generated: %s", out_path.name)

    except Exception as exc:
        logger.error("XLSX generation failed: %s", exc)
        return {"status": "error", "error": str(exc)}

    provenance = {
        "request_id": request_id,
        "session_id": session_id,
        "timestamp_utc": ts.isoformat(),
        "evidence_ids": [ev.get("doc_id") for ev in evidence],
        "label": label,
        "model_role": model_role,
        "compiler": COMPILER_VERSION,
        "output_file": safe_name,
    }
    return {
        "status": "ok",
        "file_path": str(out_path),
        "file_name": safe_name,
        "provenance": provenance,
    }
